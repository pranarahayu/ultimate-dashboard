import sys
import pandas as pd
import numpy as np
import streamlit as st
import io

import openpyxl, yattag
from openpyxl import load_workbook
from yattag import Doc, indent

st.set_page_config(page_title='Excel-to-XML Converter', layout='wide')
st.markdown('# Excel-to-XML Converter')

from menu import menu
menu()

sys.path.append("listfungsi.py")
from listfungsi import res_data
from listfungsi import cleandata
from listfungsi import cleandataver2
from listfungsi import cleandataver3
from listfungsi import converter

with st.expander("CARA PAKAI."):
    st.write("1. Upload file timeline ke file uploader pertama; 2. Download as excel, upload excel ke file uploader kedua; 3. Download file XML")

bb = st.selectbox('Pilih babak.', ['Babak 1', 'Babak 2'])
col1, col2, col3= st.columns(3)
with col1:
    fase = st.selectbox('Select Level',['1 (Satu)', '2 (Dua)', '3 (Tiga)'], key='98')
    
with col2:
    tl_data = st.file_uploader("Upload file timeline excel!")
    t1 = st.text_input('Video dimulai dari?')
    xt = converter(t1)
    try:
        tl = pd.read_excel(tl_data, skiprows=[0])
        if (fase=='1 (Satu)'):
            c_data = cleandata(tl, xt, bb)
        elif (fase=='2 (Dua)'):
            c_data = cleandataver2(tl, xt, bb)
        else:
            c_data = cleandataver3(tl, xt, bb)
        
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
            c_data.to_excel(writer, sheet_name='Sheet1', index=False)
        download = st.download_button(
            label="Download data as Excel",
            data=buffer.getvalue(),
            file_name='clean-data_'+bb+'.xlsx',
            mime='application/vnd.ms-excel',
            key = 0)
            
    except ValueError:
        st.error("Please upload the timeline file")

with col3:
    cl_data = st.file_uploader("Upload file clean-data.xlsx!")
    temp = pd.read_excel(cl_data)
    max = len(temp)
    try:
        wb = load_workbook(cl_data)
        ws = wb.worksheets[0]
        doc, tag, text = Doc().tagtext()

        with tag('file'):
          doc.asis('<!--Generated by Lapangbola-->')
          with tag('ALL_INSTANCES'):
            for row in ws.iter_rows(min_row=2, max_row=max+1, min_col=1, max_col=12):
                row = [cell.value for cell in row]
                with tag("instance"):
                    with tag("ID"):
                        text(row[0])
                    with tag("code"):
                        text(row[1])
                    with tag("start"):
                        text(row[2])
                    with tag("end"):
                        text(row[3])
                    with tag('label'):
                      with tag('group'):
                        text(row[4])
                      with tag('text'):
                        text(row[5])
                    with tag('label'):
                      with tag('group'):
                        text(row[6])
                      with tag('text'):
                        text(row[7])
                    with tag('label'):
                      with tag('group'):
                        text(row[8])
                      with tag('text'):
                        text(row[9])
                    with tag('label'):
                      with tag('group'):
                        text(row[10])
                      with tag('text'):
                        text(row[11])

        result = indent(doc.getvalue(), indentation = '    ',indent_text = True)
        with open("xml-data.xml", "w") as f:
            download1 = st.download_button(
                label="Download XML data",
                data=result,
                file_name='xml-data_'+bb+'.xml',
                mime='text/csv')

    except ValueError:
        st.error("Please upload the clean-data")
